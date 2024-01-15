import random
import math
import openpyxl


def get_value_from_cell_reference(ws, cell_reference):
    return ws[cell_reference].value

def print_statement(i, liked_recipe, custom_user, liked, type):
    if type == "SQL":
        if liked:
            return f"('{i}', UNHEX(REPLACE('{liked_recipe}', '-', '')), UNHEX(REPLACE('{custom_user}', '-', ''))) "
        else:
            return f"(UNHEX(REPLACE('{liked_recipe}', '-', '')), UNHEX(REPLACE('{custom_user}', '-', ''))) "
    if type == "Cypher":
        return "{{recipeId: '{}', userId: '{}'}} ".format(liked_recipe, custom_user)

def generate_sql_inserts_from_excel(excel_file, liked_recipe_file, custom_user_file, output_file, liked, type):
    print("Generating sql insert: \n")
    wb = openpyxl.load_workbook(excel_file)

    start = "INSERT INTO `recipe-forum`.`custom_users_liked_recipes` (`id`, `liked_recipe_id`, `custom_user_id`) VALUES " if liked else "INSERT INTO `recipe-forum`.`custom_users_disliked_recipes` (`disliked_recipe_id`, `custom_user_id`) VALUES "
    out = open(output_file, 'w')
    out.write(start)

    with open(liked_recipe_file, 'r') as lr, open(custom_user_file, 'r') as cu:
        for i, (liked_recipe_cell, custom_user_cell) in enumerate(zip(lr, cu), start=1):
            print(i)
            liked_recipe_sheet, liked_recipe_cell = liked_recipe_cell.strip().lstrip('=').split('!')
            custom_user_sheet, custom_user_cell = custom_user_cell.strip().lstrip('=').split('!')

            liked_recipe = get_value_from_cell_reference(wb[liked_recipe_sheet], liked_recipe_cell)
            custom_user = get_value_from_cell_reference(wb[custom_user_sheet], custom_user_cell)

            if liked_recipe and custom_user:  # Skip if any value is None
                out.write(print_statement(i, liked_recipe, custom_user, liked, type))

            if i % 10000 == 0:
                out.write("; \n")
                out.write(start)
            else:
                out.write(", ")

def generate_cypher_inserts_from_excel(excel_file, liked_recipe_file, custom_user_file, output_file, liked, type):
    print("Generating cypher insert: \n")
    wb = openpyxl.load_workbook(excel_file)

    start = "WITH ["
    endLiked = "] AS data UNWIND data AS row MATCH (r:Recipe {id: row.recipeId}), (u:CustomUser {id: row.userId}) CREATE (u)-[:LIKED]->(r); \n"
    endDisliked = "] AS data UNWIND data AS row MATCH (r:Recipe {id: row.recipeId}), (u:CustomUser {id: row.userId}) CREATE (u)-[:DISLIKED]->(r); \n"
    out = open(output_file, 'w')
    out.write(start)

    with open(liked_recipe_file, 'r') as lr, open(custom_user_file, 'r') as cu:
        for i, (liked_recipe_cell, custom_user_cell) in enumerate(zip(lr, cu), start=1):
            print(i)
            liked_recipe_sheet, liked_recipe_cell = liked_recipe_cell.strip().lstrip('=').split('!')
            custom_user_sheet, custom_user_cell = custom_user_cell.strip().lstrip('=').split('!')

            liked_recipe = get_value_from_cell_reference(wb[liked_recipe_sheet], liked_recipe_cell)
            custom_user = get_value_from_cell_reference(wb[custom_user_sheet], custom_user_cell)

            if liked_recipe and custom_user:  # Skip if any value is None
                out.write(print_statement(i, liked_recipe, custom_user, liked, type))

            if i % 10000 == 0:
                if liked:
                    out.write(endLiked)
                else:
                    out.write(endDisliked)
                out.write(start)
            else:
                out.write(", ")

def print_sql_relations(user_size, category_size, recipe_size, comment_size, product_category_size, product_size, line_item_size, order_size):

    user_size += 1
    category_size += 1
    recipe_size += 1
    comment_size += 1
    product_category_size += 1
    product_size += 1
    line_item_size += 1
    order_size += 1

    category = "=Category!D"
    user = "=User!J"
    recipe = "=Recipe!L"
    productCategory = "=ProductCategory!D"
    product = "=Product!G"
    order = "=Order!E"

    recipe_category = ''
    recipe_user = ''
    comment_recipe = ''
    comment_user = ''
    product_product_category = ''
    line_item_product = ''
    line_item_order = ''
    order_user = ''
    liked_recipe_liked_recipe = ''
    liked_recipe_custom_user = ''
    disliked_recipe_disliked_recipe = ''
    disliked_recipe_custom_user = ''

    print("Generating Recipe relations")

    for i in range(recipe_size):
        recipe_category += category + str(random.randint(2, category_size)) + "\n"
        recipe_user += user + str(random.randint(2, user_size)) + "\n"

    print("Generating Comment relations")

    for i in range(comment_size):
        comment_recipe += recipe + str(random.randint(2, recipe_size)) + "\n"
        comment_user += user + str(random.randint(2, user_size)) + "\n"

    print("Generating Product relations")

    for i in range(product_size):
        product_product_category += productCategory + str(random.randint(2, product_category_size)) + "\n"

    print("Generating Line Item relations")
    for i in range(line_item_size):
        line_item_product += product + str(random.randint(2, product_size)) + "\n"

    assignments = {}
    line_items = list(range(2, line_item_size + 1))
    for i in range(2, order_size + 1):
        popped_items = [line_items.pop(0) for _ in range(random.randint(1, 2))]
        assignments[i] = popped_items
    if line_items:
        for i in range(2, order_size + 1):
            if line_items.__len__() == 0:
                break
            popped_items = [line_items.pop(0) for _ in range(1)]
            assignments[random.randint(2, order_size)] += popped_items
    if line_items:
        for i in range(2, order_size + 1):
            if line_items.__len__() == 0:
                break
            popped_items = [line_items.pop(0) for _ in range(1)]
            assignments[random.randint(2, order_size)] += popped_items

    for key in assignments:
        for i in range(assignments[key].__len__()):
            line_item_order += order + str(key) + "\n"

    print("Generating Order relations")

    for i in range(order_size):
        order_user += user + str(random.randint(2, user_size)) + "\n"

    print("Writing files")

    recipe_category_file = open("sql/recipeCategory.txt", "w")
    recipe_category_file.write(recipe_category)
    recipe_category_file.close()

    recipe_user_file = open("sql/recipeUser.txt", "w")
    recipe_user_file.write(recipe_user)
    recipe_user_file.close()

    comment_recipe_file = open("sql/commentRecipe.txt", "w")
    comment_recipe_file.write(comment_recipe)
    comment_recipe_file.close()

    comment_user_file = open("sql/commentUser.txt", "w")
    comment_user_file.write(comment_user)
    comment_user_file.close()

    product_product_category_file = open("sql/productProductCategory.txt", "w")
    product_product_category_file.write(product_product_category)
    product_product_category_file.close()

    line_item_product_file = open("sql/lineItemProduct.txt", "w")
    line_item_product_file.write(line_item_product)
    line_item_product_file.close()

    line_item_order_file = open("sql/lineItemOrder.txt", "w")
    line_item_order_file.write(line_item_order)
    line_item_order_file.close()

    order_user_file = open("sql/orderUser.txt", "w")
    order_user_file.write(order_user)
    order_user_file.close()

    print("Generating Liked and Disliked Recipes relations")

    with open("sql/likedRecipeLikedRecipe.txt", "w") as liked_recipe_liked_recipe_file, \
            open("sql/likedRecipeCustomUser.txt", "w") as liked_recipe_custom_user_file, \
            open("sql/dislikedRecipeDislikedRecipe.txt", "w") as disliked_recipe_disliked_recipe_file, \
            open("sql/dislikedRecipeCustomUser.txt", "w") as disliked_recipe_custom_user_file:

        for i in range(2, math.floor(recipe_size / 2)):
            print(f"Generating relations for {i} recipe of {recipe_size / 2}")
            no_users_liking = random.randint(0, math.floor(user_size / 4))
            liking_users = random.sample(range(2, user_size), no_users_liking)
            no_users_disliking = random.randint(0, math.floor(user_size / 12))
            disliking_users = [x for x in random.sample(range(2, user_size), no_users_disliking) if
                               x not in liking_users]

            for custom_user in liking_users:
                liked_recipe_liked_recipe_file.write(f"{recipe}{i}\n")
                liked_recipe_custom_user_file.write(f"{user}{custom_user}\n")

            for custom_user in disliking_users:
                disliked_recipe_disliked_recipe_file.write(f"{recipe}{i}\n")
                disliked_recipe_custom_user_file.write(f"{user}{custom_user}\n")


def print_cypher_relations(user_size, category_size, recipe_size, comment_size, product_category_size, product_size, line_item_size, order_size):

    user_size += 1
    category_size += 1
    recipe_size += 1
    comment_size += 1
    product_category_size += 1
    product_size += 1
    line_item_size += 1
    order_size += 1

    category = "=Category!A"
    user = "=User!A"
    recipe = "=Recipe!A"
    productCategory = "=ProductCategory!A"
    product = "=Product!A"
    order = "=Order!A"

    recipe_category = ''
    recipe_user = ''
    comment_recipe = ''
    comment_user = ''
    product_product_category = ''
    line_item_product = ''
    line_item_order = ''
    order_user = ''
    liked_recipe_liked_recipe = ''
    liked_recipe_custom_user = ''
    disliked_recipe_disliked_recipe = ''
    disliked_recipe_custom_user = ''

    print("Writing files")

    sql_recipe_category_file = open("sql/recipeCategory.txt", "r")
    recipe_category = sql_recipe_category_file.read()
    sql_recipe_category_file.close()

    recipe_category = recipe_category.replace("!D", "!A")

    recipe_category_file = open("cypher/recipeCategory.txt", "w")
    recipe_category_file.write(recipe_category)
    recipe_category_file.close()

    sql_recipe_user_file = open("sql/recipeUser.txt", "r")
    recipe_user = sql_recipe_user_file.read()
    sql_recipe_user_file.close()

    recipe_user = recipe_user.replace("!J", "!A")

    recipe_user_file = open("cypher/recipeUser.txt", "w")
    recipe_user_file.write(recipe_user)
    recipe_user_file.close()

    sql_comment_recipe_file = open("sql/commentRecipe.txt", "r")
    comment_recipe = sql_comment_recipe_file.read()
    sql_comment_recipe_file.close()

    comment_recipe = comment_recipe.replace("!L", "!A")

    comment_recipe_file = open("cypher/commentRecipe.txt", "w")
    comment_recipe_file.write(comment_recipe)
    comment_recipe_file.close()

    sql_comment_user_file = open("sql/commentUser.txt", "r")
    comment_user = sql_comment_user_file.read()
    sql_comment_user_file.close()

    comment_user = comment_user.replace("!J", "!A")

    comment_user_file = open("cypher/commentUser.txt", "w")
    comment_user_file.write(comment_user)
    comment_user_file.close()

    sql_product_product_category_file = open("sql/productProductCategory.txt", "r")
    product_product_category = sql_product_product_category_file.read()
    sql_product_product_category_file.close()

    product_product_category = product_product_category.replace("!D", "!A")

    product_product_category_file = open("cypher/productProductCategory.txt", "w")
    product_product_category_file.write(product_product_category)
    product_product_category_file.close()

    sql_line_item_product_file = open("sql/lineItemProduct.txt", "r")
    line_item_product = sql_line_item_product_file.read()
    sql_line_item_product_file.close()

    line_item_product = line_item_product.replace("!G", "!A")

    line_item_product_file = open("cypher/lineItemProduct.txt", "w")
    line_item_product_file.write(line_item_product)
    line_item_product_file.close()

    sql_line_item_order_file = open("sql/lineItemOrder.txt", "r")
    line_item_order = sql_line_item_order_file.read()
    sql_line_item_order_file.close()

    line_item_order = line_item_order.replace("!E", "!A")

    line_item_order_file = open("cypher/lineItemOrder.txt", "w")
    line_item_order_file.write(line_item_order)
    line_item_order_file.close()

    sql_order_user_file = open("sql/orderUser.txt", "r")
    order_user = sql_order_user_file.read()
    sql_order_user_file.close()

    order_user = order_user.replace("!J", "!A")

    order_user_file = open("cypher/orderUser.txt", "w")
    order_user_file.write(order_user)
    order_user_file.close()

    sql_liked_recipe_liked_recipe_file = open("sql/likedRecipeLikedRecipe.txt", "r")
    liked_recipe_liked_recipe = sql_liked_recipe_liked_recipe_file.read()
    sql_liked_recipe_liked_recipe_file.close()

    liked_recipe_liked_recipe = liked_recipe_liked_recipe.replace("!L", "!A")

    liked_recipe_liked_recipe_file = open("cypher/likedRecipeLikedRecipe.txt", "w")
    liked_recipe_liked_recipe_file.write(liked_recipe_liked_recipe)
    liked_recipe_liked_recipe_file.close()

    sql_liked_recipe_custom_user_file = open("sql/likedRecipeCustomUser.txt", "r")
    liked_recipe_custom_user = sql_liked_recipe_custom_user_file.read()
    sql_liked_recipe_custom_user_file.close()

    liked_recipe_custom_user = liked_recipe_custom_user.replace("!J", "!A")

    liked_recipe_custom_user_file = open("cypher/likedRecipeCustomUser.txt", "w")
    liked_recipe_custom_user_file.write(liked_recipe_custom_user)
    liked_recipe_custom_user_file.close()

    sql_disliked_recipe_disliked_recipe_file = open("sql/dislikedRecipeDislikedRecipe.txt", "r")
    disliked_recipe_disliked_recipe = sql_disliked_recipe_disliked_recipe_file.read()
    sql_disliked_recipe_disliked_recipe_file.close()

    disliked_recipe_disliked_recipe = disliked_recipe_disliked_recipe.replace("!L", "!A")

    disliked_recipe_disliked_recipe_file = open("cypher/dislikedRecipeDislikedRecipe.txt", "w")
    disliked_recipe_disliked_recipe_file.write(disliked_recipe_disliked_recipe)
    disliked_recipe_disliked_recipe_file.close()

    sql_disliked_recipe_custom_user_file = open("sql/dislikedRecipeCustomUser.txt", "r")
    disliked_recipe_custom_user = sql_disliked_recipe_custom_user_file.read()
    sql_disliked_recipe_custom_user_file.close()

    disliked_recipe_custom_user = disliked_recipe_custom_user.replace("!J", "!A")

    disliked_recipe_custom_user_file = open("cypher/dislikedRecipeCustomUser.txt", "w")
    disliked_recipe_custom_user_file.write(disliked_recipe_custom_user)
    disliked_recipe_custom_user_file.close()


# Press the green button in the gutter to run the script.
if __name__ == '__main__':

    user_size = 2500
    category_size = 40
    recipe_size = 25000
    comment_size = 125000
    product_category_size = 15
    product_size = 100
    line_item_size = 1000
    order_size = 500

    # print_sql_relations(user_size, category_size, recipe_size, comment_size, product_category_size, product_size, line_item_size, order_size)
    # print_cypher_relations(user_size, category_size, recipe_size, comment_size, product_category_size, product_size, line_item_size, order_size)

    # generate_sql_inserts_from_excel('SQL.xlsx', 'sql/likedRecipeLikedRecipe.txt', 'sql/likedRecipeCustomUser.txt',
    #                                 'sql/insert_liked_statements.sql', True, "SQL")
    generate_sql_inserts_from_excel('SQL.xlsx', 'sql/dislikedRecipeDislikedRecipe.txt',
                                    'sql/dislikedRecipeCustomUser.txt',
                                    'sql/insert_disliked_statements.sql', False, "SQL")
    #
    # generate_cypher_inserts_from_excel('Cypher25k.xlsx', 'cypher/likedRecipeLikedRecipe.txt', 'cypher/likedRecipeCustomUser.txt',
    #                                 'cypher/insert_liked_statements.cypher', True, "Cypher")
    # generate_cypher_inserts_from_excel('Cypher25k.xlsx', 'cypher/dislikedRecipeDislikedRecipe.txt',
    #                                 'cypher/dislikedRecipeCustomUser.txt',
    #                                 'cypher/insert_dislike_statements.cypher', False, "Cypher")

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
